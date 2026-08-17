"""
Ann. B3 report builder -- adapted from the validated Colab notebook version
for use in the Streamlit app. Same tested logic, just file-bytes-in /
BytesIO-out instead of file-path-in / file-path-out, and the diagnostic
print() output is captured as text so the UI can display it.
"""

import re
import contextlib
import io
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

NFW_CATEGORY_NUMBERS = {50, 51, 52, 53, 54, 55, 56}

HEADERS = [
    "Store no.", "Store Name", "Category", "Article", "Price",
    "Stock room", "External Store", "Display", "Odds", "Subs", "Claims",
    "Total pairs", "Total value", "Total odds/subs value",
    "Pairs", "Value", "Pairs", "Value", "pairs", "Value", "Variance",
]

THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
NUMFMT = '#,##0;\\-#,##0;"-"'


def _category_number(category):
    match = re.match(r"^\s*(\d+)", str(category))
    if not match:
        return None
    return int(match.group(1))


def _is_nfw(category_num):
    return category_num in NFW_CATEGORY_NUMBERS


def _find_source_sheet(file_bytes: bytes, required: set) -> str:
    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_set = {str(h).strip() for h in header if h is not None}
            if required.issubset(header_set):
                return name
    finally:
        wb.close()
    raise ValueError(f"Could not find a sheet containing columns {sorted(required)}.")


def _build_ann_b3_inner(file_bytes: bytes) -> BytesIO:
    """
    file_bytes: raw system export (.xlsx), with the standard columns
        Code, Category, Counted, and either Retail Price or Price.
        Size-level rows (and letter-suffixed size/variant codes like
        "10140410L") for the same base Code are summed automatically.
    Everything (footwear and non-footwear) is written to Stock room.
    Rows whose Category doesn't start with a number (e.g. "Freight") are
    skipped, since there's no way to place them in the 1-56 sequence or
    classify them as FW/NFW. Rows with no digits at all in the Article
    (e.g. "Gift", "Vouch001") are dropped as non-stock rows.
    Store no. / Store Name are left blank -- the raw export doesn't carry
    store info, so fill these in by hand if needed.
    """
    required_min = {"Category", "Counted"}
    sheet_name = _find_source_sheet(file_bytes, required_min)
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    if "Article" not in df.columns:
        if "Code" in df.columns:
            df["Article"] = df["Code"]
        else:
            raise ValueError("Input file needs an 'Article' or 'Code' column.")

    if "Price" not in df.columns:
        if "Retail Price" in df.columns:
            df["Price"] = df["Retail Price"]
        else:
            raise ValueError("Input file needs a 'Price' or 'Retail Price' column.")

    missing = [c for c in ["Article", "Price", "Category", "Counted"] if c not in df.columns]
    if missing:
        raise ValueError(f"Input file is missing expected column(s): {missing}")

    # Drop rows with a genuinely missing Article or Category BEFORE any
    # string conversion -- NaN can survive .astype(str) in some pandas
    # versions and then crashes sorted()/comparisons later. Gift cards,
    # vouchers, and similar non-stock rows often show up this way.
    null_mask = df["Article"].isna() | df["Category"].isna()
    if null_mask.sum() > 0:
        examples = df.loc[null_mask, ["Article", "Category"]].head(10).to_dict("records")
        print(f"[0/3] Dropping {int(null_mask.sum())} row(s) with a blank Article or Category, e.g. {examples}")
        df = df.loc[~null_mask].reset_index(drop=True)

    if len(df) == 0:
        raise ValueError("No rows left after dropping blank Article/Category rows.")

    df["Article"] = df["Article"].astype(str).str.strip()
    df["Category"] = df["Category"].astype(str).str.strip()

    total_rows_read = len(df)
    all_raw_categories = sorted(df["Category"].unique())

    # Many articles carry a trailing size/variant letter suffix baked
    # directly into the code (e.g. "10140410L", "09525530HS", "09520002CL")
    # instead of using a separate size column -- common for Socks, Bags,
    # Shoe & Footcare, Apparel Accessories, and Clearance Footwear. Extract
    # the leading digit run as the true article code so those rows merge
    # together correctly, instead of being dropped outright (which
    # previously wiped out entire categories that had real Counted values).
    extracted_article = df["Article"].str.extract(r"^(\d+)")[0]
    no_digits = extracted_article.isna()
    if no_digits.sum() > 0:
        examples = sorted(df.loc[no_digits, "Article"].unique())[:10]
        print(f"[1/3] Dropping {int(no_digits.sum())} row(s) with no numeric article/code at all (not real stock, e.g. gift cards/freight), e.g. {examples}")
        df = df.loc[~no_digits].reset_index(drop=True)
        extracted_article = extracted_article.loc[~no_digits].reset_index(drop=True)
    else:
        print("[1/3] Every article had a numeric code (with or without a size/variant suffix).")

    if len(df) == 0:
        raise ValueError("No rows left after dropping non-numeric article/code values.")

    stripped_count = int((df["Article"].reset_index(drop=True) != extracted_article.reset_index(drop=True)).sum())
    if stripped_count > 0:
        print(f"[1b/3] Stripped a trailing size/variant suffix off {stripped_count} article code(s) (e.g. \"10140410L\" -> \"10140410\") so all its sizes merge into one article row.")

    df["Article"] = extracted_article.values

    agg = df.groupby("Article", as_index=False).agg({"Category": "first", "Price": "first", "Counted": "sum"})
    agg["Store no."] = None
    agg["Store Name"] = None
    agg["CategoryNum"] = agg["Category"].apply(_category_number)

    unparsed = agg[agg["CategoryNum"].isna()]
    if len(unparsed) > 0:
        dropped_cats = sorted(unparsed["Category"].unique())
        print(f"[2/3] Skipping {len(unparsed)} article(s) in {len(dropped_cats)} non-numeric categor(y/ies): {dropped_cats}")
        agg = agg[agg["CategoryNum"].notna()].reset_index(drop=True)
    else:
        print("[2/3] Every category parsed to a number OK.")

    agg["CategoryNum"] = agg["CategoryNum"].astype(int)
    agg["IsNFW"] = agg["CategoryNum"].apply(_is_nfw)

    # --- Diagnostic: what did we actually read for Counted, per category? ---
    diag = agg.groupby("CategoryNum").agg(
        category_text_variants=("Category", lambda s: sorted(s.unique())),
        articles=("Article", "count"),
        non_zero_articles=("Counted", lambda s: int((s != 0).sum())),
        total_counted=("Counted", "sum"),
    ).reset_index().sort_values("CategoryNum")
    print("\nPer-category summary (after steps 1-2, before the non-zero filter):")
    print(diag.to_string(index=False))

    categories_fully_zero = diag.loc[diag["non_zero_articles"] == 0, "category_text_variants"].tolist()
    if categories_fully_zero:
        print(f"\nThese categories will disappear entirely (every article's Counted is 0): {categories_fully_zero}")

    all_valid_cat_nums = {_category_number(c) for c in all_raw_categories if _category_number(c) is not None}
    print(f"\n[3/3] Total rows read: {total_rows_read}. Valid category numbers seen in file: {sorted(all_valid_cat_nums)}")

    # Only keep articles with a non-zero Counted value, so the final report
    # stays short -- a category with nothing counted in it simply won't
    # appear in the output at all.
    zero_count = int((agg["Counted"] == 0).sum())
    if zero_count > 0:
        print(f"Dropping {zero_count} article(s) with a Counted value of 0.")
        agg = agg[agg["Counted"] != 0].reset_index(drop=True)

    if len(agg) == 0:
        raise ValueError("No articles left after dropping zero-Counted rows -- check that your Counted column actually has values in it.")

    final_cat_nums = sorted(int(c) for c in agg["CategoryNum"].unique())
    print(f"Categories appearing in the final report: {final_cat_nums}\n")

    agg = agg.sort_values(["CategoryNum", "Article"]).reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ann B3"

    top_label_font = Font(bold=True, color="FFFF0000")
    for col, label in [(15, "Auditor"), (17, "Manager"), (19, "Reconciliation")]:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = top_label_font
        cell.fill = YELLOW_FILL
        cell.border = Border(bottom=THIN)

    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=THIN)

    row = 3
    first_data_row = row
    category_ranges = []

    for cat_num, cat_group in agg.groupby("CategoryNum", sort=True):
        article_start = row
        is_nfw = bool(cat_group["IsNFW"].iloc[0])
        # Everything goes in Stock room -- Display isn't used. IsNFW is
        # still tracked for category sorting and the summary block below.

        for _, r in cat_group.iterrows():
            ws.cell(row=row, column=1, value=r["Store no."])
            ws.cell(row=row, column=2, value=r["Store Name"])
            ws.cell(row=row, column=3, value=r["Category"])
            ws.cell(row=row, column=4, value=r["Article"])
            ws.cell(row=row, column=5, value=r["Price"])
            ws.cell(row=row, column=6, value=int(r["Counted"]))  # Stock room

            ws.cell(row=row, column=12, value=f"=SUM(F{row}:K{row})")
            ws.cell(row=row, column=13, value=f"=L{row}*E{row}")
            ws.cell(row=row, column=14, value=f"=(I{row}+J{row})*E{row}")

            for c in range(1, 15):
                ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=14).number_format = NUMFMT

            row += 1

        article_end = row - 1
        subtotal_row = row

        ws.cell(row=subtotal_row, column=15, value=f"=SUM(L{article_start}:L{article_end})")
        ws.cell(row=subtotal_row, column=16, value=f"=SUM(M{article_start}:M{article_end})")
        ws.cell(row=subtotal_row, column=17, value=f"=O{subtotal_row}")
        ws.cell(row=subtotal_row, column=18, value=f"=P{subtotal_row}")
        ws.cell(row=subtotal_row, column=19, value=f"=O{subtotal_row}-Q{subtotal_row}")
        ws.cell(row=subtotal_row, column=20, value=f"=P{subtotal_row}-R{subtotal_row}")

        for c in range(1, 21):
            ws.cell(row=subtotal_row, column=c).fill = YELLOW_FILL
        ws.cell(row=subtotal_row, column=14).border = Border(top=THIN)
        ws.cell(row=subtotal_row, column=14).number_format = NUMFMT
        ws.cell(row=subtotal_row, column=17).number_format = NUMFMT
        ws.cell(row=subtotal_row, column=18).number_format = NUMFMT
        ws.cell(row=subtotal_row, column=19).number_format = NUMFMT
        ws.cell(row=subtotal_row, column=20).number_format = NUMFMT
        ws.cell(row=subtotal_row, column=20).border = Border(top=THIN)

        category_ranges.append((cat_num, is_nfw, article_start, article_end, subtotal_row))
        row = subtotal_row + 1

    last_data_row = row - 1

    grand_total_row = row
    for col_letter in ["F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        ws[f"{col_letter}{grand_total_row}"] = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
    row += 2

    fw_ranges = [c for c in category_ranges if not c[1]]
    nfw_ranges = [c for c in category_ranges if c[1]]

    fw_start = fw_ranges[0][2] if fw_ranges else None
    fw_end = fw_ranges[-1][4] if fw_ranges else None
    nfw_start = nfw_ranges[0][2] if nfw_ranges else None
    nfw_end = nfw_ranges[-1][4] if nfw_ranges else None

    def _section_header(r, label):
        cell = ws.cell(row=r, column=11, value=label)
        cell.font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)

    def _pairs_value_header(r):
        ws.cell(row=r, column=11, value="\xa0")
        ws.cell(row=r, column=12, value="PAIRS").font = Font(bold=True)
        ws.cell(row=r, column=13, value="VALUE").font = Font(bold=True)

    def _fmt(r):
        ws.cell(row=r, column=12).number_format = NUMFMT
        ws.cell(row=r, column=13).number_format = NUMFMT

    _section_header(row, "AUDITOR")
    row += 1
    _pairs_value_header(row)
    row += 1
    auditor_fw_row = row
    ws.cell(row=row, column=11, value="FW")
    if fw_start:
        ws.cell(row=row, column=12, value=f"=SUM(L{fw_start}:L{fw_end})")
        ws.cell(row=row, column=13, value=f"=SUM(M{fw_start}:M{fw_end})")
    _fmt(row)
    row += 1
    auditor_nfw_row = row
    ws.cell(row=row, column=11, value="NFW")
    if nfw_start:
        ws.cell(row=row, column=12, value=f"=SUM(L{nfw_start}:L{nfw_end})")
        ws.cell(row=row, column=13, value=f"=SUM(M{nfw_start}:M{nfw_end})")
    _fmt(row)
    row += 1
    auditor_total_row = row
    ws.cell(row=row, column=11, value="TOTAL")
    ws.cell(row=row, column=12, value=f"=L{auditor_fw_row}+L{auditor_nfw_row}")
    ws.cell(row=row, column=13, value=f"=M{auditor_fw_row}+M{auditor_nfw_row}")
    _fmt(row)
    row += 2

    _section_header(row, "MANAGER")
    row += 1
    _pairs_value_header(row)
    row += 1
    manager_fw_row = row
    ws.cell(row=row, column=11, value="FW")
    ws.cell(row=row, column=12, value=f"=L{auditor_fw_row}")
    ws.cell(row=row, column=13, value=f"=M{auditor_fw_row}")
    _fmt(row)
    row += 1
    manager_nfw_row = row
    ws.cell(row=row, column=11, value="NFW")
    ws.cell(row=row, column=12, value=f"=L{auditor_nfw_row}")
    ws.cell(row=row, column=13, value=f"=M{auditor_nfw_row}")
    _fmt(row)
    row += 1
    manager_total_row = row
    ws.cell(row=row, column=11, value="TOTAL")
    ws.cell(row=row, column=12, value=f"=L{auditor_total_row}")
    ws.cell(row=row, column=13, value=f"=M{auditor_total_row}")
    _fmt(row)
    row += 2

    _section_header(row, "RECONCILIATION")
    row += 1
    _pairs_value_header(row)
    row += 1
    ws.cell(row=row, column=11, value="FW")
    ws.cell(row=row, column=12, value=f"=L{manager_fw_row}-L{auditor_fw_row}")
    ws.cell(row=row, column=13, value=f"=M{manager_fw_row}-M{auditor_fw_row}")
    _fmt(row)
    row += 1
    ws.cell(row=row, column=11, value="NFW")
    ws.cell(row=row, column=12, value=f"=L{manager_nfw_row}-L{auditor_nfw_row}")
    ws.cell(row=row, column=13, value=f"=M{manager_nfw_row}-M{auditor_nfw_row}")
    _fmt(row)
    row += 1
    ws.cell(row=row, column=11, value="TOTAL")
    ws.cell(row=row, column=12, value=f"=L{manager_total_row}-L{auditor_total_row}")
    ws.cell(row=row, column=13, value=f"=M{manager_total_row}-M{auditor_total_row}")
    _fmt(row)

    widths = {"A": 8.7, "C": 20.4, "D": 8.7, "F": 8.9, "M": 11.1, "N": 11.9,
              "O": 8.7, "P": 11.5, "Q": 8.7, "R": 12.3, "S": 8.7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_ann_b3_report(file_bytes: bytes):
    """Runs the builder and captures its diagnostic print() output.

    Returns (buffer, diagnostic_text) so the UI can show what happened
    at each filtering stage.
    """
    diagnostic_stream = io.StringIO()
    with contextlib.redirect_stdout(diagnostic_stream):
        buf = _build_ann_b3_inner(file_bytes)
    return buf, diagnostic_stream.getvalue()
