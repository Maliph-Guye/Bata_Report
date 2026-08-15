"""
Core report-building logic for the Stock Report Tool.

This is plain Python (pandas + openpyxl), independent of Streamlit, so it
can be tested and reused on its own.
"""

from collections import defaultdict
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_COLUMNS = [
    "Code",
    "Retail Price",
    "SOH",
    "Counted",
    "Variance",
    "Total Retail",
    "Category",
]

THIN = Side(style="thin", color="999999")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
# Shared helpers
# --------------------------------------------------------------------------

def find_source_sheet(file_bytes: bytes, required: set) -> str:
    """Pick the sheet that actually contains the raw stock data.

    Uses read_only mode and only looks at the header row, so this stays
    fast even on very large files.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_set = {str(h).strip() for h in header if h is not None}
            if required.issubset(header_set):
                return name
    finally:
        wb.close()
    raise ValueError(
        f"Could not find a sheet containing columns {sorted(required)} in the uploaded file. "
        f"Check for typos, extra spaces, or different capitalization in your column headers."
    )


def style_header_row(ws, ncols: int):
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for cell in ws[1][:ncols]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def apply_body_font(ws, from_row: int = 2, to_row: int | None = None):
    to_row = to_row or ws.max_row
    body_font = Font(name="Arial")
    for row in ws.iter_rows(min_row=from_row, max_row=to_row):
        for cell in row:
            cell.font = body_font


def apply_money_format(ws, col_indexes: list, from_row: int, to_row: int):
    for c in col_indexes:
        col_letter = get_column_letter(c)
        for row_idx in range(from_row, to_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = "#,##0"


def auto_width(ws):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_letter].width = max_len + 3


def _ensure_columns(df: pd.DataFrame, cols_with_defaults: dict):
    for col, default in cols_with_defaults.items():
        if col not in df.columns:
            df[col] = default


def _to_buffer(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------
# 1. Inputting -- SOH-only pre-count list
# --------------------------------------------------------------------------

def build_inputting_report(file_bytes: bytes) -> BytesIO:
    sheet_name = find_source_sheet(file_bytes, {"Code", "SOH"})
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]

    _ensure_columns(df, {"Counted": 0, "Variance": 0, "Total Retail": 0, "Retail Price": 0, "Category": ""})
    df["Code"] = df["Code"].astype(str).str.strip()

    agg = df.groupby("Code", as_index=False).agg(
        {
            "SOH": "sum",
            "Counted": "sum",
            "Variance": "sum",
            "Total Retail": "sum",
            "Retail Price": "first",
            "Category": "first",
        }
    )
    agg = agg[agg["SOH"] != 0].sort_values("Code").reset_index(drop=True)
    agg = agg[OUTPUT_COLUMNS]

    wb = Workbook()
    ws = wb.active
    ws.title = "SOH For Input"
    ws.append(OUTPUT_COLUMNS)
    for row in agg.itertuples(index=False, name=None):
        ws.append(row)

    style_header_row(ws, len(OUTPUT_COLUMNS))
    apply_body_font(ws)
    apply_money_format(ws, [2, 6], 2, ws.max_row)
    auto_width(ws)
    ws.freeze_panes = "A2"

    return _to_buffer(wb)


# --------------------------------------------------------------------------
# 2. Reconciliation -- collapse to one row per code, drop inactive articles
# --------------------------------------------------------------------------

def build_reconciliation_report(file_bytes: bytes) -> BytesIO:
    sheet_name = find_source_sheet(file_bytes, {"Code", "SOH", "Counted", "Variance", "Total Retail"})
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
    df.columns = [str(c).strip() for c in df.columns]
    df["Code"] = df["Code"].astype(str).str.strip()

    agg = df.groupby("Code", as_index=False).agg(
        {
            "SOH": "sum",
            "Counted": "sum",
            "Variance": "sum",
            "Total Retail": "sum",
            "Retail Price": "first",
            "Category": "first",
        }
    )
    agg = agg[~((agg["SOH"] == 0) & (agg["Counted"] == 0))].sort_values("Code").reset_index(drop=True)
    agg = agg[OUTPUT_COLUMNS]

    wb = Workbook()
    ws = wb.active
    ws.title = "By Code"
    ws.append(OUTPUT_COLUMNS)
    for row in agg.itertuples(index=False, name=None):
        ws.append(row)

    style_header_row(ws, len(OUTPUT_COLUMNS))
    apply_body_font(ws)
    apply_money_format(ws, [2, 6], 2, ws.max_row)
    auto_width(ws)
    ws.freeze_panes = "A2"

    return _to_buffer(wb)


# --------------------------------------------------------------------------
# 3. Variance Report -- 3-sheet workbook with netting and totals
# --------------------------------------------------------------------------

def _net_offsetting_pairs(df: pd.DataFrame):
    """Remove pairs of rows whose Total Retail values are exact opposites
    (e.g. 500 and -500). Since they sum to zero, removing them never
    changes the Total Retail total.
    """
    df = df.reset_index(drop=True)
    pool = defaultdict(list)
    remove_indices = set()

    for idx, row in df.iterrows():
        key = round(float(row["Total Retail"]), 2)
        neg_key = -key
        if pool[neg_key]:
            match_idx = pool[neg_key].pop()
            remove_indices.add(idx)
            remove_indices.add(match_idx)
        else:
            pool[key].append(idx)

    kept = df.drop(index=list(remove_indices)).reset_index(drop=True)
    return kept, len(remove_indices) // 2


def build_variance_report(file_bytes: bytes):
    sheet_name = find_source_sheet(file_bytes, {"Code", "SOH", "Counted", "Retail Price"})
    raw_df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)

    df = raw_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df["Code"] = df["Code"].astype(str).str.strip()

    agg = df.groupby("Code", as_index=False).agg(
        {"SOH": "sum", "Counted": "sum", "Retail Price": "first", "Category": "first"}
    )
    agg["Variance"] = agg["Counted"] - agg["SOH"]
    agg["Total Retail"] = agg["Variance"] * agg["Retail Price"]
    agg = agg[OUTPUT_COLUMNS]

    variance_df = agg[agg["Variance"] != 0].sort_values("Code").reset_index(drop=True)
    netted_df, pairs_removed = _net_offsetting_pairs(variance_df)
    netted_df["Reasons for Variance"] = ""

    wb = Workbook()

    # --- Sheet 1: Stock -- raw copy, completely untouched, no styling ---
    stock_ws = wb.active
    stock_ws.title = "Stock"
    stock_ws.append(list(raw_df.columns))
    for row in raw_df.itertuples(index=False, name=None):
        stock_ws.append(row)

    # --- Sheet 2: Variance ---
    var_ws = wb.create_sheet("Variance")
    var_ws.append(OUTPUT_COLUMNS)
    for row in variance_df.itertuples(index=False, name=None):
        var_ws.append(row)
    style_header_row(var_ws, len(OUTPUT_COLUMNS))
    apply_body_font(var_ws)
    apply_money_format(var_ws, [2, 6], 2, var_ws.max_row)
    auto_width(var_ws)
    var_ws.freeze_panes = "A2"

    # --- Sheet 3: Variance Report ---
    rep_cols = OUTPUT_COLUMNS + ["Reasons for Variance"]
    rep_ws = wb.create_sheet("Variance Report")
    rep_ws.append(rep_cols)
    for row in netted_df.itertuples(index=False, name=None):
        rep_ws.append(row)

    n_data_rows = len(netted_df)
    last_data_row = 1 + n_data_rows

    style_header_row(rep_ws, len(rep_cols))
    apply_body_font(rep_ws, 2, last_data_row)
    apply_money_format(rep_ws, [2, 6], 2, last_data_row)

    for r in range(1, last_data_row + 1):
        for c in range(1, len(rep_cols) + 1):
            rep_ws.cell(row=r, column=c).border = THIN_BORDER

    totals_row = last_data_row + 1
    sum_headers = {"SOH", "Counted", "Variance", "Total Retail"}
    money_headers = {"Retail Price", "Total Retail"}
    total_font = Font(name="Arial", bold=True)

    for i, header in enumerate(rep_cols, start=1):
        col_letter = get_column_letter(i)
        cell = rep_ws.cell(row=totals_row, column=i)
        if header == "Code":
            cell.value = "TOTAL"
        elif header in sum_headers:
            cell.value = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
            if header in money_headers:
                cell.number_format = "#,##0"
        cell.font = total_font
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    auto_width(rep_ws)
    rep_ws.freeze_panes = "A2"

    return _to_buffer(wb), pairs_removed
